import os
from PyPDF2 import PdfReader
from docx import Document
import openpyxl
import csv
import json
import xml.etree.ElementTree as ET
import hashlib

# ANSI escape codes for colored output
GREEN = "\033[32m"
RESET = "\033[0m"

def get_md5(file_path):
    md5_hash = hashlib.md5()
    with open(file_path, "rb") as file:
        for chunk in iter(lambda: file.read(4096), b""):
            md5_hash.update(chunk)
    return md5_hash.hexdigest()

def get_sha1(file_path):
    sha1_hash = hashlib.sha1()
    with open(file_path, "rb") as file:
        for chunk in iter(lambda: file.read(4096), b""):
            sha1_hash.update(chunk)
    return sha1_hash.hexdigest()

def search_pdf_for_content(directory, content_to_find):
    pdf_files_with_content = []

    for root, _, files in os.walk(directory):
        for file in files:
            if file.lower().endswith('.pdf'):
                pdf_file = os.path.join(root, file)
                try:
                    pdf_reader = PdfReader(pdf_file)
                    for page_num in range(len(pdf_reader.pages)):
                        page = pdf_reader.pages[page_num]
                        page_text = page.extract_text()
                        if content_to_find.lower() in page_text.lower():
                            pdf_files_with_content.append(pdf_file)
                            break  # If content is found, no need to check other pages
                except Exception as e:
                    print(f"Error reading {pdf_file}: {e}")

    return pdf_files_with_content

def search_docx_for_content(directory, content_to_find):
    docx_files_with_content = []

    for root, _, files in os.walk(directory):
        for file in files:
            if file.lower().endswith('.docx'):
                docx_file = os.path.join(root, file)
                try:
                    doc = Document(docx_file)
                    for paragraph in doc.paragraphs:
                        if content_to_find.lower() in paragraph.text.lower():
                            docx_files_with_content.append(docx_file)
                            break
                except Exception as e:
                    print(f"Error reading {docx_file}: {e}")

    return docx_files_with_content

def search_text_files_for_content(directory, content_to_find):
    text_files_with_content = []

    for root, _, files in os.walk(directory):
        for file in files:
            ext = file.lower().split('.')[-1]
            if ext in ('txt', 'csv', 'log'):
                text_file = os.path.join(root, file)
                try:
                    with open(text_file, 'r', encoding='utf-8') as file:
                        text = file.read()
                        if content_to_find.lower() in text.lower():
                            text_files_with_content.append(text_file)
                except Exception as e:
                    print(f"Error reading {text_file}: {e}")

    return text_files_with_content

def search_excel_files_for_content(directory, content_to_find):
    excel_files_with_content = []

    for root, _, files in os.walk(directory):
        for file in files:
            ext = file.lower().split('.')[-1]
            if ext in ('xlsx', 'xls', 'xlsm', 'xlsb', 'xlt', 'xltx', 'xltm', 'ods'):
                excel_file = os.path.join(root, file)
                try:
                    wb = openpyxl.load_workbook(excel_file)
                    for sheet_name in wb.sheetnames:
                        sheet = wb[sheet_name]
                        for row in sheet.iter_rows():
                            for cell in row:
                                if content_to_find.lower() in str(cell.value).lower():
                                    excel_files_with_content.append(excel_file)
                                    break
                except Exception as e:
                    print(f"Error reading {excel_file}: {e}")

    return excel_files_with_content

def search_csv_files_for_content(directory, content_to_find):
    csv_files_with_content = []

    for root, _, files in os.walk(directory):
        for file in files:
            if file.lower().endswith('.csv'):
                csv_file = os.path.join(root, file)
                try:
                    with open(csv_file, 'r', encoding='utf-8') as file:
                        csv_reader = csv.reader(file)
                        for row in csv_reader:
                            for cell in row:
                                if content_to_find.lower() in cell.lower():
                                    csv_files_with_content.append(csv_file)
                                    break
                except Exception as e:
                    print(f"Error reading {csv_file}: {e}")

    return csv_files_with_content

def search_sql_files_for_content(directory, content_to_find):
    sql_files_with_content = []

    for root, _, files in os.walk(directory):
        for file in files:
            if file.lower().endswith('.sql'):
                sql_file = os.path.join(root, file)
                try:
                    with open(sql_file, 'r', encoding='utf-8') as file:
                        sql_text = file.read()
                        if content_to_find.lower() in sql_text.lower():
                            sql_files_with_content.append(sql_file)
                except Exception as e:
                    print(f"Error reading {sql_file}: {e}")

    return sql_files_with_content

def search_json_files_for_content(directory, content_to_find):
    json_files_with_content = []

    for root, _, files in os.walk(directory):
        for file in files:
            if file.lower().endswith('.json'):
                json_file = os.path.join(root, file)
                try:
                    with open(json_file, 'r', encoding='utf-8') as file:
                        json_data = json.load(file)
                        if content_to_find.lower() in str(json_data).lower():
                            json_files_with_content.append(json_file)
                except Exception as e:
                    print(f"Error reading {json_file}: {e}")

    return json_files_with_content

def search_xml_files_for_content(directory, content_to_find):
    xml_files_with_content = []

    for root, _, files in os.walk(directory):
        for file in files:
            if file.lower().endswith('.xml'):
                xml_file = os.path.join(root, file)
                try:
                    with open(xml_file, 'r', encoding='utf-8') as file:
                        xml_text = file.read()
                        if content_to_find.lower() in xml_text.lower():
                            xml_files_with_content.append(xml_file)
                except Exception as e:
                    print(f"Error reading {xml_file}: {e}")

    return xml_files_with_content

def search_code_files_for_content(directory, content_to_find):
    code_files_with_content = []

    for root, _, files in os.walk(directory):
        for file in files:
            ext = file.lower().split('.')[-1]
            if ext in ('php', 'py', 'sh', 'bat', 'rb', 'pl', 'c', 'cpp', 'jar', 'jdk', 'asp', 'html', 'js', 'css'):
                code_file = os.path.join(root, file)
                try:
                    with open(code_file, 'r', encoding='utf-8') as file:
                        code_text = file.read()
                        if content_to_find.lower() in code_text.lower():
                            code_files_with_content.append(code_file)
                except Exception as e:
                    print(f"Error reading {code_file}: {e}")

    return code_files_with_content

def search_files_by_md5(directory, md5_to_find):
    files_with_md5 = []

    for root, _, files in os.walk(directory):
        for file in files:
            file_path = os.path.join(root, file)
            try:
                md5 = get_md5(file_path)
                if md5_to_find.lower() == md5.lower():
                    files_with_md5.append(file_path)
            except Exception as e:
                print(f"Error reading {file_path}: {e}")

    return files_with_md5

def search_files_by_sha1(directory, sha1_to_find):
    files_with_sha1 = []

    for root, _, files in os.walk(directory):
        for file in files:
            file_path = os.path.join(root, file)
            try:
                sha1 = get_sha1(file_path)
                if sha1_to_find.lower() == sha1.lower():
                    files_with_sha1.append(file_path)
            except Exception as e:
                print(f"Error reading {file_path}: {e}")

    return files_with_sha1

def main():
    search_directory = input("Enter the directory to search for files: ")

    while True:
        print("\nMenu:")
        print("1. Search in PDF files")
        print("2. Search in DOCX files")
        print("3. Search in TXT, CSV, and LOG files")
        print("4. Search in Excel files")
        print("5. Search in CSV files")
        print("6. Search in SQL files")
        print("7. Search in JSON files")
        print("8. Search in XML files")
        print("9. Search in PHP, Python, Shell, Batch, Ruby, Perl, C/C++, JAR, JDK, ASP, HTML, JS, CSS, and Code files")
        print("10. Search files by MD5 checksum")
        print("11. Search files by SHA1 checksum")
        print("12. Exit")

        choice = input("Enter your choice (1/2/3/4/5/6/7/8/9/10/11/12): ")

        if choice == '1':
            content_to_find = input("Enter the content to search for: ")
            found_files = search_pdf_for_content(search_directory, content_to_find)
            file_type = "PDF files"
        elif choice == '2':
            content_to_find = input("Enter the content to search for: ")
            found_files = search_docx_for_content(search_directory, content_to_find)
            file_type = "DOCX files"
        elif choice == '3':
            content_to_find = input("Enter the content to search for: ")
            found_files = search_text_files_for_content(search_directory, content_to_find)
            file_type = "TXT, CSV, and LOG files"
        elif choice == '4':
            content_to_find = input("Enter the content to search for: ")
            found_files = search_excel_files_for_content(search_directory, content_to_find)
            file_type = "Excel files"
        elif choice == '5':
            content_to_find = input("Enter the content to search for: ")
            found_files = search_csv_files_for_content(search_directory, content_to_find)
            file_type = "CSV files"
        elif choice == '6':
            content_to_find = input("Enter the content to search for: ")
            found_files = search_sql_files_for_content(search_directory, content_to_find)
            file_type = "SQL files"
        elif choice == '7':
            content_to_find = input("Enter the content to search for: ")
            found_files = search_json_files_for_content(search_directory, content_to_find)
            file_type = "JSON files"
        elif choice == '8':
            content_to_find = input("Enter the content to search for: ")
            found_files = search_xml_files_for_content(search_directory, content_to_find)
            file_type = "XML files"
        elif choice == '9':
            content_to_find = input("Enter the content to search for: ")
            found_files = search_code_files_for_content(search_directory, content_to_find)
            file_type = "Code files"
        elif choice == '10':
            md5_to_find = input("Enter the MD5 checksum to search for: ")
            found_files = search_files_by_md5(search_directory, md5_to_find)
            file_type = "files with MD5 checksum"
        elif choice == '11':
            sha1_to_find = input("Enter the SHA1 checksum to search for: ")
            found_files = search_files_by_sha1(search_directory, sha1_to_find)
            file_type = "files with SHA1 checksum"
        elif choice == '12':
            break
        else:
            print("Invalid choice. Please enter a valid option.")
            continue

        if found_files:
            print(f"{file_type} ({len(found_files)} files found):")
            for idx, file in enumerate(found_files, start=1):
                print(f"{idx}. {GREEN}{file}{RESET}")
        else:
            print(f"No {file_type} found.")

if __name__ == "__main__":
    main()
